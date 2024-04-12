import torch.utils.data as data
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import pandas as pd
import csv
from openpyxl import load_workbook

class MultiDatasetFolder(data.Dataset):
    def __init__(self, root, index, transform):
        self.dataset = []
        for r in root:
            if r.endswith('PIQ23/'):
                self.dataset.append(PIQ23Folder(r, index, transform))
                # self.dataset.extend([PIQ23Folder(r, index, transform) for _ in range(len(r))])
            elif r.endswith('SPAQ/'):
                self.dataset.append(SPAQFolder(r, None, transform))
            elif r.endswith('LIVEW/'):
                self.dataset.append(LIVEChallengeFolder(r, None, transform))
            elif r.endswith('koniq10k/'):
                self.dataset.append(Koniq_10kFolder(r, None, transform))
            elif r.endswith('BID/'):
                self.dataset.append(BIDFolder(r, None, transform))
            else:
                raise ValueError(f"Unknown dataset: {r}")
            
        self.dataset = data.ConcatDataset(self.dataset)
        self.scene = []
        for d in self.dataset.datasets:
            self.scene += d.get_scene_list()
    
    def __getitem__(self, index):
        # try:
        return self.dataset.__getitem__(index)
        # except Exception as e:
        #     print(f'Error: {e}')
        #     return self.dataset.__getitem__(0)
    
    def __len__(self):
        return len(self.dataset)
    
    def get_scene_list(self):
        return self.scene
    

class PIQ23Folder(data.Dataset):
    def __init__(self, root, index:list, transform, scene_base=0):
        data_dir = root
        all_set = os.path.join(data_dir, 'Scores_Overall.csv')
        all_df = pd.read_csv(all_set)
        imgs_name = all_df['IMAGE PATH'].tolist()
        imgpath = [os.path.join(data_dir, img_name.replace('\\', '/')) for img_name in imgs_name]
        labels = all_df['JOD'].tolist()
        scene = [scene_base+int(s.split("_")[-1]) for s in all_df['SCENE'].tolist()]

        sample = []

        for i, item in enumerate(index):
            sample.append(dict(
                path = imgpath[item],
                target = labels[item],
                scene = scene[item],
                img_name = imgs_name[item],
            ))
        self.samples = sample
        self.transform = transform

        # # fit all scene target to base_scene
        # base_scene = self.samples[0]['scene']
        # base_scene_target = [sample['target'] for sample in self.samples if sample['scene'] == base_scene]
        # # compute a, b for each scene
        # scene_ab_dict = {}
        # for s in set(scene):
        #     scene_target = [sample['target'] for sample in self.samples if sample['scene'] == s]
        #     a, b = np.polyfit(scene_target, base_scene_target, 1)
        #     scene_ab_dict[s] = (a, b)
        # # ajust target
        # for sample in self.samples:
        #     a, b = scene_ab_dict[sample['scene']]
        #     sample['target'] = a * sample['target'] + b

        # # normalize target for each scene
        # if is_training:
        #     scene_std_mean_dict = {}
        #     for s in set(scene):
        #         scene_target = [sample['target'] for sample in self.samples if sample['scene'] == s]
        #         scene_std_mean_dict[s] = (np.std(scene_target), np.mean(scene_target))
        #     for sample in self.samples:
        #         std, mean = scene_std_mean_dict[sample['scene']]
        #         sample['target'] = (sample['target'] - mean) / std
        


    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        sample = self.samples[index]
        path = sample['path']
        target = sample['target']
        scene = sample['scene']
        img_name = sample['img_name']
        
        img = pil_loader(path)
        if self.transform is not None:
            img, img_pt = (tf(img) for tf in self.transform)

        return dict(
            img=img,
            img_pt=img_pt,
            label=target,
            scene=scene,
            img_name=img_name
        )

    def __len__(self):
        length = len(self.samples)
        return length
    
    def get_scene_list(self):
        return [sample['scene'] for sample in self.samples]

class SPAQFolder(PIQ23Folder):
    def __init__(self, root, index=None, transform=None, scene_base=100):
        # data_dir = os.path.join(root, 'SPAQ')
        data_dir = root
        all_set = os.path.join(data_dir, 'annotations/MOS and Image attribute scores.xlsx')
        all_df = pd.read_excel(all_set)
        imgs_name = all_df['Image name'].tolist()
        imgpath = [os.path.join(data_dir, f'TestImage/{img_name}') for img_name in imgs_name]
        labels = all_df['MOS'].tolist()
        # scene = ['spaq_0'] * len(imgs_name)
        scene = [scene_base] * len(imgs_name)
        if index is None:
            index = list(range(len(imgs_name)))

        self.samples = []
        for i, item in enumerate(index):
            self.samples.append(dict(
                path = imgpath[item],
                target = labels[item],
                scene = scene[item],
                img_name = imgs_name[item],
            ))
        self.transform = transform
    
    

class LIVEFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = scipy.io.loadmat(os.path.join(root, 'dmos_realigned.mat'))
        labels = dmos['dmos_new'].astype(np.float32)

        orgs = dmos['orgs']
        refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refnames_all = refnames_all['refnames_all']

        sample = []

        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((imgpath[item], labels[0][item]))
                # print(self.imgpath[item])
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = '%s%s%s' % ('img', str(index), '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename


class LIVEChallengeFolder(PIQ23Folder):

    def __init__(self, root, index=None, transform=None, scene_base=200):

        # data_dir = os.path.join(root, 'LIVEW')
        data_dir = root
        imgpath = scipy.io.loadmat(os.path.join(data_dir, 'Data', 'AllImages_release.mat'))
        imgpath = imgpath['AllImages_release']
        imgpath = imgpath[7:1169]
        mos = scipy.io.loadmat(os.path.join(data_dir, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]
        # scene = ['livew_0'] * len(imgpath)
        scene = [scene_base] * len(imgpath)

        if index is None:
            index = list(range(len(imgpath)))
        self.samples = []
        for i, item in enumerate(index):
            self.samples.append(dict(
                path = os.path.join(data_dir, 'Images', imgpath[item][0][0]),
                target = labels[item],
                scene = scene[item],
                img_name = imgpath[item][0][0],
            ))

        self.transform = transform



class CSIQFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath,'.png')
        txtpath = os.path.join(root, 'csiq_label.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[0]))
            target.append(words[1])
            ref_temp = words[0].split(".")
            refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []

        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((os.path.join(root, 'dst_imgs_all', imgnames[item]), labels[item]))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class Koniq_10kFolder(PIQ23Folder):

    def __init__(self, root, index=None, transform=None, scene_base=300):
        # data_dir = os.path.join(root, 'koniq10k')
        data_dir = root
        imgname = []
        mos_all = []
        csv_file = os.path.join(data_dir, 'koniq10k_scores_and_distributions.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = float(row['MOS_zscore'])
                mos_all.append(mos)
        scene = [scene_base] * len(imgname)
        if index is None:
            index = list(range(len(imgname)))

        sample = []
        for i, item in enumerate(index):
            sample.append(dict(
                path = os.path.join(data_dir, '1024x768', imgname[item]),
                target = mos_all[item],
                scene = scene[item],
                img_name = imgname[item],
            ))

        self.samples = sample
        self.transform = transform


class BIDFolder(PIQ23Folder):

    def __init__(self, root, index, transform, scene_base=400):
        data_dir = root
        imgname = []
        mos_all = []

        xls_file = os.path.join(data_dir, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            # mos = np.array(mos)
            # mos = mos.astype(np.float32)
            mos_all.append(mos)
            if count == 587:
                break
        # scene = ['bid_0'] * len(imgname)
        scene = [scene_base] * len(imgname)
        if index is None:
            index = list(range(len(imgname)))

        sample = []
        for i, item in enumerate(index):
            sample.append(dict(
                path = os.path.join(data_dir, imgname[item]),
                target = mos_all[item],
                scene = scene[item],
                img_name = imgname[item],
            ))

        self.samples = sample
        self.transform = transform

class TID2013Folder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        refpath = os.path.join(root, 'reference_images')
        refname = getTIDFileName(refpath,'.bmp.BMP')
        txtpath = os.path.join(root, 'mos_with_names.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refnames_all.append(ref_temp[0][1:])
        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((os.path.join(root, 'distorted_images', imgnames[item]), labels[item]))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename


def pil_loader(path):
    with open(path, 'rb') as f:
        img = Image.open(f)
        return img.convert('RGB')