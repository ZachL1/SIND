import os
from openpyxl import load_workbook

from g_iqa.g_datasets.folder.PIQ23_base import PIQ23Folder


class BIDFolder(PIQ23Folder):

    def __init__(self, root, index, transform, scene_base=400):
        data_dir = root
        imgname = []
        mos_all = []

        xls_file = os.path.join(data_dir, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            # mos = np.array(mos)
            # mos = mos.astype(np.float32)
            mos_all.append(mos)
            if count == 587:
                break
        # scene = ['bid_0'] * len(imgname)
        scene = [scene_base] * len(imgname)
        if index is None:
            index = list(range(len(imgname)))

        sample = []
        for i, item in enumerate(index):
            sample.append(dict(
                path = os.path.join(data_dir, imgname[item]),
                target = mos_all[item],
                scene = scene[item],
                img_name = imgname[item],
            ))

        self.samples = sample
        self.transform = transform