import os
import json
import pandas as pd
import numpy as np
import scipy
import random
from PIL import Image


base_dir = '/home/dzc/workspace/G-IQA'
data_dir = f'{base_dir}/data'


def get_imgnames_from_json(json_file):
  with open(json_file, 'r') as f:
    imgnames = json.load(f)
  return set(item['image'] if 'image' in item else item['img_path'] for item in imgnames)



############################ PIQ23 ############################
def piq23_generator(domain_id_base = 0):
  all_df = pd.read_csv(os.path.join(data_dir, 'PIQ23', 'Scores_Overall.csv'))
  imgs_name = all_df['IMAGE PATH'].tolist()
  labels = all_df['JOD'].tolist()
  scenes = all_df['SCENE'].tolist()

  all_files = []
  for img_name, label, scene in zip(imgs_name, labels, scenes):
    all_files.append(dict(
      image = f"PIQ23/{img_name}".replace('\\', '/'),
      score = label,
      scene = scene,
      domain_id = domain_id_base + int(scene.split("_")[-1]),
    ))
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))

  with open(f'{base_dir}/data_json/all/piq23_all.json', 'w') as f:
    json.dump({
      'files': all_files,
      'domain_name': {domain_id_base + int(scene.split("_")[-1]): scene for scene in sorted(set(scenes))},
      }, f, indent=2)
  print(f'PIQ23 all: {len(all_files)}')



############################ SPAQ ############################

def extract_positive_indices(row):
  return [idx for idx, value in enumerate(row) if value > 0]

def spaq_generator(domain_id_base = 100):

  all_df = pd.read_excel(f'{data_dir}/SPAQ/annotations/MOS and Image attribute scores.xlsx')
  scene_df = pd.read_excel(f'{data_dir}/SPAQ/annotations/Scene category labels.xlsx')
  assert all_df['Image name'].tolist() == scene_df['Image name'].tolist()
  imgs_name = all_df['Image name'].tolist()
  labels = all_df['MOS'].tolist()

  # 由于SPAQ中一个图片可能对应多个场景，因此这里提取所有场景类别作为candidate，每轮epoch从对应的candidate中随机挑选一个作为其场景标签
  # scene 标签只用于训练时区分不同场景（场景采样），对于测试而言没有任何意义
  scene_candidate = scene_df.drop(columns='Image name').apply(lambda row: extract_positive_indices(row), axis=1).tolist()
  # Animal	Cityscape	Human	Indoor scene	Landscape	Night scene	Plant	Still-life	Others
  scene_dict = {
    0: 'Animal',
    1: 'Cityscape',
    2: 'Human',
    3: 'Indoor scene',
    4: 'Landscape',
    5: 'Night scene',
    6: 'Plant',
    7: 'Still-life',
    8: 'Others',
  }

  all_files = []
  for img_name, label, scene_list in zip(imgs_name, labels, scene_candidate):
    all_files.append(dict(
      image = f'SPAQ/TestImage/{img_name}',
      score = label,
      scene = [scene_dict[scene] for scene in scene_list],
      domain_id = [domain_id_base + scene for scene in scene_list],
    ))

    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))
  
  # random select for multi-label scene
  for item in all_files:
    item['scene'] = random.choice(item['scene'])
    item['domain_id'] = random.choice(item['domain_id'])
  # or CLIP select
  # from clip_based_scene_classify import classify_images
  # all_files = classify_images(all_files, data_dir=data_dir)

  with open(f'{base_dir}/data_json/all/spaq_all.json', 'w') as f:
    json.dump({
      'files': all_files,
      'domain_name': {domain_id_base + scene: scene_dict[scene] for scene in scene_dict.keys()},
      }, f, indent=2)


  # split train/test follow by q-align
  train_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/training_sft/train_spaq.json')
  test_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/test_jsons/test_spaq.json')
  assert len(train_img & test_img) == 0
  train_files = [item for item in all_files if item['image'].replace('SPAQ/TestImage', 'spaq') in train_img]
  test_files = [item for item in all_files if item['image'].replace('SPAQ/TestImage', 'spaq') in test_img]
  assert len(train_files) + len(test_files) == len(train_img) + len(test_img)
  with open(f'{base_dir}/data_json/for_cross_set/train/spaq_train.json', 'w') as f:
    json.dump({'files': train_files}, f, indent=2)
  with open(f'{base_dir}/data_json/for_cross_set/test/spaq_test.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'SPAQ all: {len(all_files)}, train: {len(train_files)}, test: {len(test_files)}')


############################ LIVE Challenge ############################
def livec_generator(domain_id_base = 200):
  imgpath = scipy.io.loadmat(os.path.join(data_dir, 'LIVEChallenge', 'Data', 'AllImages_release.mat'))
  imgpath = imgpath['AllImages_release']
  imgpath = imgpath
  mos = scipy.io.loadmat(os.path.join(data_dir, 'LIVEChallenge', 'Data', 'AllMOS_release.mat'))
  labels = mos['AllMOS_release']
  labels = labels[0]

  all_files = []
  for img, label in zip(imgpath, labels):
    all_files.append(dict(
      image = f'LIVEChallenge/Images/{img[0][0]}',
      score = label,
      domain_id = domain_id_base,
    ))
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))
  
  with open(f'{base_dir}/data_json/all/livec_all.json', 'w') as f:
    json.dump({'files': all_files}, f, indent=2)
  
  # split train/test follow by q-align
  # all for test
  test_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/test_jsons/livec.json')
  test_files = [item for item in all_files if item['image'].replace('LIVEChallenge/Images', 'livec') in test_img]
  assert len(test_files) == len(test_img)
  with open(f'{base_dir}/data_json/for_cross_set/test/livec.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'LIVE Challenge all: {len(all_files)}, test: {len(test_files)}')


############################ Koniq_10k ############################
def koniq_generator(domain_id_base = 300):
  all_df = pd.read_csv(os.path.join(data_dir, 'koniq10k/koniq10k_scores_and_distributions.csv'))
  imgs_name = all_df['image_name'].tolist()
  labels = all_df['MOS_zscore'].tolist()

  all_files = []
  for img_name, label in zip(imgs_name, labels):
    all_files.append(dict(
      image = f'koniq10k/1024x768/{img_name}',
      score = label,
      domain_id = domain_id_base,
    ))
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))

  # CLIP scene classify
  from clip_based_scene_classify import classify_images
  all_files = classify_images(all_files, data_dir=data_dir)

  with open(f'{base_dir}/data_json/all/koniq10k_all.json', 'w') as f:
    json.dump({'files': all_files}, f, indent=2)

  # split train/test follow by q-align
  train_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/training_sft/train_koniq.json')
  test_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/test_jsons/test_koniq.json')
  assert len(train_img & test_img) == 0
  train_files = [item for item in all_files if item['image'].replace('koniq10k/1024x768', 'koniq') in train_img]
  test_files = [item for item in all_files if item['image'].replace('koniq10k/1024x768', 'koniq') in test_img]
  assert len(train_files) + len(test_files) == len(train_img) + len(test_img)
  with open(f'{base_dir}/data_json/for_cross_set/train/koniq10k_train.json', 'w') as f:
    json.dump({'files': train_files}, f, indent=2)
  with open(f'{base_dir}/data_json/for_cross_set/test/koniq10k_test.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'Koniq all: {len(all_files)}, train: {len(train_files)}, test: {len(test_files)}')


############################ BID ############################
def bid_generator(domain_id_base = 400):
  from openpyxl import load_workbook
  workbook = load_workbook(os.path.join(data_dir, 'BID/DatabaseGrades.xlsx'))
  booksheet = workbook.active
   
  count = 1
  all_files = []
  for row in booksheet.rows:
    count += 1
    img_num = (booksheet.cell(row=count, column=1).value)
    img_name = f"DatabaseImage{img_num:04d}.JPG"
    mos = (booksheet.cell(row=count, column=2).value)

    all_files.append(dict(
      image = f'BID/{img_name}',
      score = mos,
      domain_id = domain_id_base,
    ))
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))
    if count == 587:
      break

  with open(f'{base_dir}/data_json/all/bid_all.json', 'w') as f:
    json.dump({'files': all_files}, f, indent=2)

  # all for test
  with open(f'{base_dir}/data_json/for_cross_set/test/bid.json', 'w') as f:
    json.dump({'files': all_files}, f, indent=2)
  print(f'BID all: {len(all_files)}')

############################ CID2013 ############################
def cid2013_generator(domain_id_base = 450):
  all_df = pd.read_excel(os.path.join(data_dir, 'CID2013/CID2013 data - version 12112014.xlsx'), sheet_name='CID2013 MOS')
  imgs_name = all_df['Source_ID'].tolist()
  mos = all_df['Realigned MOS'].tolist()

  all_files = []
  for img_name, label in zip(imgs_name, mos):
    all_files.append(dict(
      image = f'CID2013/{img_name}.jpg',
      score = label,
      domain_id = domain_id_base,
    ))
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))

  # all for test
  with open(f'{base_dir}/data_json/for_cross_set/test/cid2013.json', 'w') as f:
    json.dump({'files': all_files}, f, indent=2)
  print(f'CID2013 all: {len(all_files)}')




############################ AGIQA-3K ############################
def agiqa3k_generator(domain_id_base = 500):
  all_df = pd.read_csv(os.path.join(data_dir, 'AGIQA-3K/data.csv'))
  imgs_name = all_df['name'].tolist()
  labels = all_df['mos_quality'].tolist()
  style = all_df['style'].tolist()
  style = ['other' if s is np.nan else s for s in style]
  style_dict = {s: i for i, s in enumerate(sorted(set(style)))}

  all_files = []
  for img_name, label, style in zip(imgs_name, labels, style):
    all_files.append(dict(
      image = f'AGIQA-3K/images/{img_name}',
      score = label,
      style = style,
      domain_id = domain_id_base + style_dict[style],
    ))
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))

  # split test follow by q-align
  test_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/test_jsons/agi.json')
  test_files = [item for item in all_files if item['image'].replace('AGIQA-3K/images', 'agi-cgi') in test_img]
  assert len(test_files) == len(test_img)
  with open(f'{base_dir}/data_json/for_cross_set/test/agiqa3k.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'AGIQA-3K all: {len(all_files)}, test: {len(test_files)}')


############################ KADID-10k ############################
def kadid10k_generator(domain_id_base = 600):
  all_df = pd.read_csv(os.path.join(data_dir, 'kadid10k/dmos.csv'))
  imgs_name = all_df['dist_img'].tolist()
  labels = all_df['dmos'].tolist()
  dist_types = [img.split('_')[1] for img in imgs_name]

  all_files = []
  for img_name, label, dist_type in zip(imgs_name, labels, dist_types):
    all_files.append(dict(
      image = f'kadid10k/images/{img_name}',
      score = label,
      dist_type = dist_type,
      domain_id = domain_id_base + int(dist_type),
    ))
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))
  
  with open(f'{base_dir}/data_json/all/kadid10k_all.json', 'w') as f:
    json.dump({
      'files': all_files,
      'domain_name': {domain_id_base + int(dist_type): dist_type for dist_type in sorted(set(dist_types))},
      }, f, indent=2)

  # split train/test follow by q-align
  train_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/training_sft/train_kadid.json')
  test_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/test_jsons/test_kadid.json')
  # assert len(train_img & test_img) == 0
  train_files = [item for item in all_files if item['image'].replace('kadid10k/images', 'kadid10k') in train_img]
  test_files = [item for item in all_files if item['image'].replace('kadid10k/images', 'kadid10k') in test_img]
  assert len(train_files) + len(test_files) == len(train_img) + len(test_img)
  with open(f'{base_dir}/data_json/for_cross_set/train/kadid10k_train.json', 'w') as f:
    json.dump({'files': train_files}, f, indent=2)
  with open(f'{base_dir}/data_json/for_cross_set/test/kadid10k_test.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'KADID-10k all: {len(all_files)}, train: {len(train_files)}, test: {len(test_files)}')
  
############################ LIVE ############################
def getFileName(path, suffix):
  filename = []
  f_list = os.listdir(path)
  for i in f_list:
      if os.path.splitext(i)[1] == suffix:
          filename.append(i)
  return filename
  
def live_generator(domain_id_base = 700):
  ''' Modified from https://github.com/SSL92/hyperIQA '''
  def getDistortionTypeFileName(path, num):
    filename = []
    index = 1
    for i in range(0, num):
        name = '%s%s%s' % ('img', str(index), '.bmp')
        filename.append(os.path.join(path, name))
        index = index + 1
    return filename
  
  root = os.path.join(data_dir, 'LIVE')
  refpath = os.path.join(root, 'refimgs')
  refname = getFileName(refpath, '.bmp')

  jp2kroot = os.path.join(root, 'jp2k')
  jp2kname = getDistortionTypeFileName(jp2kroot, 227)

  jpegroot = os.path.join(root, 'jpeg')
  jpegname = getDistortionTypeFileName(jpegroot, 233)

  wnroot = os.path.join(root, 'wn')
  wnname = getDistortionTypeFileName(wnroot, 174)

  gblurroot = os.path.join(root, 'gblur')
  gblurname = getDistortionTypeFileName(gblurroot, 174)

  fastfadingroot = os.path.join(root, 'fastfading')
  fastfadingname = getDistortionTypeFileName(fastfadingroot, 174)

  imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

  dmos = scipy.io.loadmat(os.path.join(root, 'dmos_realigned.mat'))
  labels = dmos['dmos_new']

  orgs = dmos['orgs']
  refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
  refnames_all = refnames_all['refnames_all']

  dist_type_dict = {
    'jp2k': 0,
    'jpeg': 1,
    'wn': 2,
    'gblur': 3,
    'fastfading': 4,
  }

  all_files = []
  for i, rname in enumerate(refname):
    train_sel = (rname == refnames_all)
    # train_sel = train_sel * ~orgs.astype(np.bool_)
    train_sel = np.where(train_sel == True)
    train_sel = train_sel[1].tolist()
    for j, item in enumerate(train_sel):
      dist_type = imgpath[item].split('/')[-2]
      all_files.append(dict(
        image = imgpath[item].replace(f'{data_dir}/', ''),
        score = labels[0][item],
        dist_type = dist_type,
        domain_id = domain_id_base + dist_type_dict[dist_type],
      ))
      assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))

  with open(f'{base_dir}/data_json/all/live_all.json', 'w') as f:
    json.dump({
      'files': all_files,
      'domain_name': {domain_id_base + dist_type_dict[dist_type]: dist_type for dist_type in dist_type_dict.keys()},
      }, f, indent=2)

  # split test follow by q-align
  test_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/test_jsons/live.json')
  test_files = [item for item in all_files if item['image'].replace('LIVE', 'live') in test_img]
  assert len(test_files) == len(test_img) # TODO: not equal
  with open(f'{base_dir}/data_json/for_cross_set/test/live.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'LIVE all: {len(all_files)}, test: {len(test_files)}')
  

############################ CSIQ ############################
def csiq_generator(domain_id_base = 800):
  ''' Modified from https://github.com/SSL92/hyperIQA '''
  refpath = os.path.join(data_dir, 'CSIQ', 'src_imgs')
  refname = getFileName(refpath,'.png')
  txtpath = os.path.join(data_dir, 'CSIQ', 'csiq_label.txt') # see: https://github.com/SSL92/hyperIQA/blob/master/csiq_label.txt
  fh = open(txtpath, 'r')
  imgnames = []
  target = []
  refnames_all = []
  for line in fh:
      line = line.split('\n')
      words = line[0].split()
      imgnames.append((words[0]))
      target.append(words[1])
      ref_temp = words[0].split(".")
      refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

  labels = np.array(target)
  refnames_all = np.array(refnames_all)

  dist_type_dict = {
    'awgn': 0,
    'blur': 1,
    'contrast': 2,
    'fnoise': 3,
    'jpeg': 4,
    'jpeg2000': 5,
  }

  all_files = []
  for i, item in enumerate(refname):
    train_sel = (refname[i] == refnames_all)
    train_sel = np.where(train_sel == True)
    train_sel = train_sel[0].tolist()
    for j, item in enumerate(train_sel):
      dist_type = imgnames[item].split('.')[1].lower()
      all_files.append(dict(
        image = os.path.join('CSIQ', 'dst_imgs', dist_type, imgnames[item]),
        score = float(labels[item]),
        dist_type = dist_type,
        domain_id = domain_id_base + dist_type_dict[dist_type],
      ))

  with open(f'{base_dir}/data_json/all/csiq_all.json', 'w') as f:
    json.dump({
      'files': all_files,
      'domain_name': {domain_id_base + dist_type_dict[dist_type]: dist_type for dist_type in dist_type_dict.keys()},
      }, f, indent=2)

  # split test follow by q-align
  test_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/test_jsons/csiq.json')
  test_files = [item for item in all_files if item['image'].replace('CSIQ', 'csiq') in test_img]
  # assert len(test_files) == len(test_img) # TODO: miss gt in csiq_label.txt
  with open(f'{base_dir}/data_json/for_cross_set/test/csiq.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'CSIQ all: {len(all_files)}, test: {len(test_files)}')


############################ TID2013 ############################
def tid2013_generator(domain_id_base = 900):
  mos_with_names = pd.read_csv(f'{data_dir}/TID2013/mos_with_names.txt', sep=' ', header=None, index_col=1)

  all_files = []
  for img_name, score in mos_with_names[0].items():
    # dist_dype = int(img_name.split('_')[1])
    dist_dype = int(img_name.split('_')[0][1:])
    all_files.append(dict(
      image = f'TID2013/distorted_images/{img_name}',
      score = score,
      dist_dype = dist_dype,
      domain_id = domain_id_base + dist_dype,
    ))
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))
  
  with open(f'{base_dir}/data_json/all/tid2013_all.json', 'w') as f:
    json.dump({
      'files': all_files,
      'domain_name': {domain_id_base + dist_dype: dist_dype for dist_dype in set([item['dist_dype'] for item in all_files])},
      }, f, indent=2)
  
  # split train/test by image.split('_')[0].lower()
  all_img = set(item.split('_')[0].lower() for item in mos_with_names[0].keys())
  train_img = set(random.sample(all_img, k=int(len(all_img)*0.8)))
  test_img = all_img - train_img
  train_files = [item for item in all_files if item['image'].replace('TID2013/distorted_images/', '').split('_')[0].lower() in train_img]
  test_files = [item for item in all_files if item['image'].replace('TID2013/distorted_images/', '').split('_')[0].lower() in test_img]
  assert len(train_files) + len(test_files) == len(all_files)
  with open(f'{base_dir}/data_json/for_cross_set/train/tid2013_train.json', 'w') as f:
    json.dump({'files': train_files}, f, indent=2)
  with open(f'{base_dir}/data_json/for_cross_set/test/tid2013_test.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'TID2013 train: {len(train_files)}, test: {len(test_files)}')



############################ EVA ############################
def eva_generator(domain_id_base = 1100):
  
  # animals, architectures and city scenes, human, natural and rural scenes, still life, other
  cat_dict = {
    "1": "animals",
    "2": "architectures and city scenes",
    "3": "human",
    "4": "natural and rural scenes",
    "5": "still life",
    "6": "other",
  }

  all_files = []
  with open(f'{data_dir}/EVA/eva-dataset-master/data/image_content_category.csv', 'r') as f:
    lines = f.readlines()

  info = pd.read_csv(f'{data_dir}/EVA/eva-dataset-master/data/votes_filtered.csv', sep='=', dtype={'image_id': str})
  for line in lines[1:]:
    id, cat = line.strip().split(',')
    id, cat = id.strip('"'), cat.strip('"')
    score = info['score'][info['image_id']==id].to_numpy()
    if len(score) == 0:
      continue
    all_files.append({
      'image': f'EVA/eva-dataset-master/images/EVA_category/EVA_category/{cat}/{id}.jpg',
      'score': np.mean(score),
      'scene': cat_dict[cat],
      'domain_id': int(cat) + domain_id_base,
    })
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))

  with open(f'{base_dir}/data_json/all/eva_all.json', 'w') as f:
    json.dump({
      'files': all_files,
      'domain_name': {int(cat) + domain_id_base: cat_dict[cat] for cat in cat_dict.keys()},
      }, f, indent=2)
  print(f'EVA all: {len(all_files)}')


############################ PARA ############################

def para_generator(domain_id_base = 1200):
  '''
  # old version
  def __init__(self, root, index=None, transform=None, scene_base=1000, iaa=True):
      data_dir = root
      all_set = os.path.join(data_dir, 'annotation/PARA-GiaaAll.csv')
      all_df = pd.read_csv(all_set)
      imgs_name = all_df['imageName'].tolist()
      sessions = all_df['sessionId'].tolist()
      imgpath = [os.path.join(data_dir, f'imgs/{session}/{img_name}') for session, img_name in zip(sessions, imgs_name)]
      labels = all_df['aestheticScore_mean'].tolist() if iaa else all_df['qualityScore_mean'].tolist()
      
      # domain label to unique int
      all_scene = sorted(all_df['semantic'].unique())
      scene_dict = {scene: i for i, scene in enumerate(all_scene)}
      scene = [scene_dict[s] + scene_base for s in all_df['semantic'].tolist()]

      if index is None:
          index = list(range(len(imgs_name)))

      # check domain category
      print('domain category:', set(scene[i] for i in index))
      
      self.samples = []
      for i, item in enumerate(index):
          self.samples.append(dict(
              path = imgpath[item],
              target = labels[item],
              scene = scene[item],
              img_name = imgs_name[item],
          ))
      self.transform = transform
  '''
  annos_data = pd.read_csv(f'{data_dir}/PARA/annotation/PARA-GiaaAll.csv')
  all_scene = sorted(annos_data['semantic'].unique())
  scene_dict = {scene: i for i, scene in enumerate(all_scene)}

  all_files = []
  for _, row in annos_data.iterrows():
    all_files.append({
      'image': f'PARA/imgs/{row["sessionId"]}/{row["imageName"]}',
      'score': row['aestheticScore_mean'],
      'scene': row['semantic'],
      'domain_id': domain_id_base + scene_dict[row['semantic']],
    })
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image'])), os.path.join(data_dir, all_files[-1]['image'])
  
  with open(f'{base_dir}/data_json/all/para_all.json', 'w') as f:
    json.dump({
      'files': all_files,
      'domain_name': {domain_id_base + scene_dict[scene]: scene for scene in all_scene},
      }, f, indent=2)
  print(f'PARA all: {len(all_files)}')

  # split train/test follow official split
  test_csv = f'{data_dir}/PARA/annotation/PARA-GiaaTest.csv'
  train_csv = f'{data_dir}/PARA/annotation/PARA-GiaaTrain.csv'
  # get sessionId/imageName, image_names = PARA/imgs/{sessionId}/{imageName}
  test_df = pd.read_csv(test_csv)
  train_df = pd.read_csv(train_csv)
  test_img = [f'PARA/imgs/{row["sessionId"]}/{row["imageName"]}' for _, row in test_df.iterrows()]
  train_img = [f'PARA/imgs/{row["sessionId"]}/{row["imageName"]}' for _, row in train_df.iterrows()]
  test_files = [item for item in all_files if item['image'] in test_img]
  train_files = [item for item in all_files if item['image'] in train_img]
  assert len(test_files) == len(test_img)
  assert len(train_files) == len(train_img)
  with open(f'{base_dir}/data_json/for_cross_set/train/para_train.json', 'w') as f:
    json.dump({'files': train_files}, f, indent=2)
  with open(f'{base_dir}/data_json/for_cross_set/test/para_test.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'PARA train: {len(train_files)}, test: {len(test_files)}')


############################ AVA ############################
def ava_generator(domain_id_base = 1300):
  def get_mean(ratings:list):
    sum_rat = 0
    for rat in range(1,11):
      sum_rat += ratings[rat-1] * rat
    return sum_rat / sum(ratings)

  def read_tag(file_path:str) -> dict:
    tags = {}
    with open(file_path, 'r') as f:
      for line in f:
        line = line.strip()
        spl = line.find(' ')
        tag_id, tag = line[:spl], line[spl+1:]
        tags[int(tag_id)] = tag
    return tags

  def check_file(file_path):
    try:
      Image.open(file_path).convert('RGB')
    except:
      return False
    return True

  # read semantic tags and meta data
  tags = read_tag(os.path.join(data_dir, 'AVA/data/tags.txt'))
  meta_data = pd.read_csv(os.path.join(data_dir, 'AVA/data/AVA.txt'), sep=' ', index_col=1, header=None, dtype={1:str})

  all_files = []
  for img_id in meta_data.index:
    if not check_file(os.path.join(data_dir, f'AVA/data/image/{img_id}.jpg')):
      print(os.path.join(data_dir, f'AVA/data/image/{img_id}.jpg'))
      continue # check if image is valid

    meta = meta_data.loc[img_id,:].to_numpy().tolist()
    ratings = meta[1:11]
    sem_tags = [tags[tag] for tag in meta[11:13] if tag != 0]
    all_files.append({
      'image': f'AVA/data/image/{img_id}.jpg',
      'score': get_mean(ratings),
      'ratings': ratings,
      'sem_tags': sem_tags,
      'domain_id': domain_id_base,
    })
    assert os.path.exists(os.path.join(data_dir, all_files[-1]['image'])), os.path.join(data_dir, all_files[-1]['image'])

  with open(f'{base_dir}/data_json/all/ava_all.json', 'w') as f:
    json.dump({'files': all_files}, f, indent=2)

  # split train/test follow by q-align
  train_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/training_sft/train_ava.json')
  test_img = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/test_jsons/test_ava.json')
  assert len(train_img & test_img) == 0
  train_files = [item for item in all_files if item['image'].replace('AVA/data/image', 'ava_images') in train_img]
  test_files = [item for item in all_files if item['image'].replace('AVA/data/image', 'ava_images') in test_img]
  # assert len(train_files) + len(test_files) == len(train_img) + len(test_img)
  print(len(train_files), len(test_files), len(train_img), len(test_img))
  with open(f'{base_dir}/data_json/for_cross_set/train/ava_train.json', 'w') as f:
    json.dump({'files': train_files}, f, indent=2)
  with open(f'{base_dir}/data_json/for_cross_set/test/ava_test.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  print(f'AVA all: {len(all_files)}, train: {len(train_files)}, test: {len(test_files)}')


############################ NNID ############################
def nnid_generator(domain_id_base = 1400):
  mos_file = {
    'sub512': f'{data_dir}/NNID/mos512_with_names.txt',
    'sub1024': f'{data_dir}/NNID/mos1024_with_names.txt',
    'Sub2048': f'{data_dir}/NNID/mos2048_with_names.txt',
  }
  all_files = []
  for i, (sub, file) in enumerate(mos_file.items()):
    with open(file, 'r') as f:
      for line in f:
        mos, img_name = line.strip().split()
        all_files.append({
          'image': f'NNID/{sub}/{img_name}',
          'score': float(mos),
          'domain_id': domain_id_base + i,
        })
        assert os.path.exists(os.path.join(data_dir, all_files[-1]['image']))

  with open(f'{base_dir}/data_json/all/nnid_all.json', 'w') as f:
    json.dump({'files': all_files}, f, indent=2)
  with open(f'{base_dir}/data_json/for_cross_set/test/nnid.json', 'w') as f:
    json.dump({'files': all_files}, f, indent=2)
  print(f'NNID all: {len(all_files)}')


############################ PIPAL ############################
def pipal_generator(domain_id_base = 1500):
  import glob

  ref_names = os.listdir(f'{data_dir}/PIPAL/training/Train_Ref')
  # select_names = [name.split('.')[0] for name in random.sample(ref_names, k=20)]
  select_names = [name.split('.')[0] for name in ref_names]

  image_names = []
  image_scores = []
  for name in select_names:
    mos_file = f'{data_dir}/PIPAL/training/Train_Label/{name}.txt'
    with open(mos_file, 'r') as f:
      for line in f:
        img_name, score = line.strip().split(',')
        image_names.append(img_name)
        image_scores.append(float(score))
  
  train_files = []
  for img_name, score in zip(image_names, image_scores):
    img_path = glob.glob(f'{data_dir}/PIPAL/training/Distortion_*/{img_name}')
    assert len(img_path) == 1, img_path
    train_files.append({
      'image': img_path[0].replace(data_dir+'/', ''),
      'score': (score - 868.2988) / (1857 - 868.2988),
      'domain_id': domain_id_base + int(img_name.split('_')[0][1:]),
    })
    assert os.path.exists(os.path.join(data_dir, train_files[-1]['image']))

  with open(f'{base_dir}/data_json/for_cross_set/train/pipal_train.json', 'w') as f:
    json.dump({'files': train_files}, f, indent=2)
  
  image_names = []
  image_scores = []
  with open(f'{data_dir}/PIPAL/validation/val_label.txt', 'r') as f:
    for line in f:
      img_name, score = line.strip().split(',')
      image_names.append(img_name)
      image_scores.append(float(score))
  test_files = []
  for img_name, score in zip(image_names, image_scores):
    img_path = glob.glob(f'{data_dir}/PIPAL/validation/Dis/{img_name}')
    assert len(img_path) == 1, img_path
    test_files.append({
      'image': img_path[0].replace(data_dir+'/', ''),
      'score': score,
      'domain_id': domain_id_base + int(img_name.split('_')[0][1:]),
    })
    assert os.path.exists(os.path.join(data_dir, test_files[-1]['image']))
  
  with open(f'{base_dir}/data_json/for_cross_set/test/pipal_test.json', 'w') as f:
    json.dump({'files': test_files}, f, indent=2)
  
  with open(f'{base_dir}/data_json/all/pipal_all.json', 'w') as f:
    json.dump({'files': train_files + test_files}, f, indent=2)
  print(f'PIPAL train: {len(train_files)}, test: {len(test_files)}')




def check_koniq_spaq_kadia():
  koniq = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/training_sft/train_koniq.json')
  spaq = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/training_sft/train_spaq.json')
  kadid = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/training_sft/train_kadid.json')

  koniq_spaq_kadid = get_imgnames_from_json(f'{data_dir}/Q-Align/playground/data/training_sft/train_koniq_spaq_kadid.json')

  print(len(koniq), len(spaq), len(kadid), len(koniq_spaq_kadid))
  print(len(koniq & koniq_spaq_kadid), len(spaq & koniq_spaq_kadid), len(kadid & koniq_spaq_kadid))
  

if __name__ == '__main__':
  # piq23_generator()
  # spaq_generator()
  # livec_generator()
  # koniq_generator()
  # bid_generator()
  # cid2013_generator()
  # agiqa3k_generator()
  # kadid10k_generator()
  # live_generator()
  # csiq_generator()
  tid2013_generator()
  # eva_generator()
  # para_generator()
  # ava_generator()
  # nnid_generator()
  # pipal_generator()

  # check_koniq_spaq_kadia()